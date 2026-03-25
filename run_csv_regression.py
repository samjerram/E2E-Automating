#!/usr/bin/env python3
"""
CSV Regression Pack: Upload a CSV, run automation for each row, get a summary.

End-to-end flow:
  1. Customer uploads CSV
  2. Automation runs for each row (behind the scenes)
  3. Summary CSV is given at the end

Two display modes:
  --quiet    Behind the scenes: no browser visible, minimal output, summary only at end
  --summary  Same as quiet (summary always shown); use --verbose to see per-preset progress

Usage:
  python3 run_csv_regression.py
    → Prompts for CSV path (user-friendly; paste or drag file path)

  python3 run_csv_regression.py my_regression.csv
    → Runs directly with the given CSV

  python3 run_csv_regression.py my_regression.csv --quiet
    → Behind the scenes, summary at end (good for customers)

  python3 run_csv_regression.py my_regression.csv --verbose
    → Show progress for each preset (for debugging)

CSV format: See CSV_UPLOAD_FORMAT.md
Example: example_regression_upload.csv
"""
import argparse
import csv
import json
import os
import re
import subprocess
import sys
import tempfile
import time
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from config import get_portal_base_url, get_neos_internal_creds

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

RUN_PRESET_SCRIPT = Path(__file__).resolve().parent / "run_preset.py"
PRESETS_DIR = Path(__file__).resolve().parent / "presets"
RESULTS_DIR = Path(__file__).resolve().parent / "p2nni_regression_results"
MAX_ROWS = 999  # No practical limit; was 5
# When running in-process (Flask/UI), stop waiting after this many seconds so the run never hangs forever.
PRESET_TIMEOUT_SEC = 300  # 5 minutes per preset


@dataclass
class RowResult:
    row: int
    preset_id: str
    postcode: str
    exit_code: int
    duration_sec: float
    result: str  # "Pass" or "Fail"
    error_detail: str = ""  # Stderr snippet when failed
    # Power BI / SharePoint bridge
    order_id: str = ""
    quotation_id: str = ""  # PE/DP: digits from "PE-28620239" (same value for both columns)
    line_id: str = ""      # Quotation line ID (which broadband tile); empty if not scraped
    tcv_total: str = ""    # Total Contract Value from quote page (e.g. 16216.20)
    company_name: str = ""
    start_port_speed: str = ""  # Bearer / B-End Port (e.g. 10000, 1000, 100)
    start_supplier: str = ""   # Provider from green-highlighted price tile (e.g. Virgin, Neos via Openreach)
    install_price: str = ""    # Up-front install cost from quote page (e.g. 4072.00 or 0.00)
    annual_rental: str = ""    # Annual Charge from order page (e.g. 6874.67)
    ftpp_aggregation: str = "" # FTTP Aggregation from order page (e.g. 0.00)
    add_on: str = ""          # Add-ons e.g. "fttp included; shadowvlan surcharged"
    order_number: str = ""    # Order reference e.g. "O-12f21ca8" (for summary + hyperlink)
    quote_number: str = ""    # Quote reference e.g. "Q-ffd8feb9" (for summary + hyperlink)
    order_url: str = ""       # Full URL to order page (for hyperlink in CSV/Excel)
    quote_url: str = ""       # Full URL to quote page (for hyperlink in CSV/Excel)
    end_port_speed: str = ""
    path_speed: str = ""  # Bandwidth
    term_length: str = ""
    completed_at: str = ""  # Per-row completion timestamp
    basket_id: str = ""  # Demo 2: internal user flow
    negotiated_annual: str = ""
    negotiated_install: str = ""
    list_annual: str = ""
    list_install: str = ""


def find_presets(presets_dir: Path) -> list[Path]:
    if not presets_dir.exists():
        return []
    presets = []
    for p in presets_dir.rglob("*.json"):
        if "_old" in p.parts:
            continue
        presets.append(p)
    presets.sort(key=lambda x: str(x).lower())
    return presets


def resolve_preset(spec: str) -> Path | None:
    """Resolve preset path from id or relative path."""
    spec = str(spec).strip()
    if not spec:
        return None
    presets = find_presets(PRESETS_DIR)
    # Direct path
    p = Path(spec)
    if p.is_file():
        return p
    if not p.suffix:
        p = p.with_suffix(".json")
    candidate = PRESETS_DIR / p
    if candidate.is_file():
        return candidate
    candidate = (PRESETS_DIR / "p2nni" / p).resolve()
    if candidate.is_file():
        return candidate
    # Match by id or partial name
    spec_lower = spec.lower()
    for preset in presets:
        if spec_lower in str(preset).lower():
            return preset
        try:
            data = json.loads(preset.read_text(encoding="utf-8"))
            if data.get("id", "").lower() == spec_lower:
                return preset
        except Exception:
            pass
    return None


def _get_cell(row: dict, key_lower: str) -> str:
    """Get cell value case-insensitively."""
    for k, v in row.items():
        if k and str(k).strip().lower() == key_lower:
            return (v or "").strip()
    return ""


def _mbps_to_bearer(mbps: str) -> str:
    m = str(mbps).strip()
    if not m:
        return "1 Gbps"
    v = int(m) if m.isdigit() else 1000
    if v >= 10000:
        return "10 Gbps"
    if v >= 1000:
        return "1 Gbps"
    return "100 Mbps"


def _parse_bearer(val: str) -> str:
    """Parse bearer from CSV: '10 Gbps', '100 Gbps', '1 Gbps', '10000', etc."""
    v = str(val).strip().lower()
    if not v:
        return "1 Gbps"
    if "100" in v and ("g" in v or "gbps" in v) and "10 " not in v and "1000" not in v:
        return "100 Gbps"
    if "10" in v and ("g" in v or "gbps" in v):
        return "10 Gbps"
    if "1" in v and ("g" in v or "gbps" in v):
        return "1 Gbps"
    if "100" in v and "m" in v:
        return "100 Mbps"
    return _mbps_to_bearer(val)


def _parse_bandwidth(val: str) -> str:
    """Parse bandwidth from CSV: '1 Gbps', '10 Gbps', '100 Gbps', '1000', etc."""
    v = str(val).strip().lower()
    if not v:
        return "1 Gbps"
    if "100" in v and ("g" in v or "gbps" in v) and "10 " not in v and "1000" not in v:
        return "100 Gbps"
    if "10" in v and ("g" in v or "gbps" in v):
        return "10 Gbps"
    if "1" in v and ("g" in v or "gbps" in v) and "10" not in v:
        return "1 Gbps"
    if v.isdigit():
        n = int(v)
        if n >= 10000:
            return "10 Gbps" if n < 100000 else "100 Gbps"
        if n >= 1000:
            return "1 Gbps"
        return f"{n} Mbps"
    return _mbps_to_bearer(val)  # fallback


def _parse_bearer(val: str) -> str:
    """Parse bearer from '10 Gbps', '10000', etc. into normalized form."""
    v = str(val or "").strip().lower().replace(" ", "")
    if not v:
        return "1 Gbps"
    if "100g" in v or "100000" in v:
        return "100 Gbps"
    if "10g" in v or "10000" in v:
        return "10 Gbps"
    if "1g" in v or "1000" in v:
        return "1 Gbps"
    if "100" in v and "g" not in v:
        return "100 Mbps"
    return _mbps_to_bearer(val)


def _parse_bandwidth(val: str) -> str:
    """Parse bandwidth from '1 Gbps', '1000', etc. into normalized form."""
    raw = str(val or "").strip()
    v = raw.lower().replace(" ", "")
    if not v:
        return "1 Gbps"
    if "100g" in v or "100000" in v:
        return "100 Gbps"
    if "10g" in v or "10000" in v:
        return "10 Gbps"
    if "1g" in v or v == "1000":
        return "1 Gbps"
    if v in ("500", "200", "100"):
        return f"{v} Mbps"
    if v.isdigit():
        n = int(v)
        if n >= 1000:
            return "1 Gbps" if n < 10000 else "10 Gbps"
        return f"{n} Mbps"
    return "1 Gbps"


def _months_to_term(months: str) -> str:
    m = str(months).strip()
    if not m:
        return "3 years"
    try:
        v = int(float(m))  # Handles "24", " 24", "24.0" from Excel
    except (ValueError, TypeError):
        v = 36
    return f"{v} months" if v not in (12, 24, 36, 48, 60) else {
        12: "1 year", 24: "2 years", 36: "3 years", 48: "4 years", 60: "5 years"
    }.get(v, f"{v} months")


def _yes_no(val: str, default: bool = False) -> bool:
    v = str(val).strip().lower()
    if not v:
        return default
    return v in ("yes", "y", "1", "true")


def _cell(row: dict, csv_cols: list[str], defaults: dict, internal_key: str, default: str = "") -> str:
    """Get cell: try CSV columns, then defaults[internal_key], else default."""
    for col in csv_cols:
        v = _get_cell(row, col)
        if v:
            return v
    return str(defaults.get(internal_key, "") or default)


def _build_preset_from_row(row: dict, billing_defaults: dict, row_index: int) -> dict:
    """Build preset dict from full-format CSV row. Billing defaults from row 1."""
    base = json.loads((PRESETS_DIR / "p2nni" / "p2nni_1g_1g_36m_upfront_shadowvlan_LC_AC_LR_VlanNo_0-48h_Hazards_AutoNegYes.json").read_text(encoding="utf-8"))
    base["base_url"] = get_portal_base_url()
    b_end = _get_cell(row, "b-end postcode") or _get_cell(row, "b_end postcode") or _get_cell(row, "postcode") or billing_defaults.get("b_end_postcode", "SP2 8NJ")
    bearer_raw = _get_cell(row, "bearer") or _get_cell(row, "b-end port (mbps)") or _get_cell(row, "b_end port (mbps)") or "1000"
    bearer_b = _parse_bearer(bearer_raw)
    bandwidth_raw = _get_cell(row, "bandwidth") or _get_cell(row, "bandwidth (mbps)") or "1000"
    bandwidth_str = _parse_bandwidth(bandwidth_raw)
    preferred_supplier = (_get_cell(row, "preferred supplier") or _get_cell(row, "service provider") or "").strip()
    if preferred_supplier.lower().startswith("no specified"):
        preferred_supplier = ""
    preset_id = f"csv_row_{row_index}"
    base["id"] = preset_id
    term_months = _get_cell(row, "term length (months)") or _get_cell(row, "term length")
    base["quote"] = {
        "product_tile": "Ethernet Point to NNI",
        "b_end_postcode": b_end,
        "select_first_choose_location": False,
        "select_first_nni": True,
        # Optional: if provided, try to pick this broadband supplier's price tile; otherwise keep default
        "preferred_supplier": preferred_supplier,
        "shadow_vlan_required": _yes_no(_get_cell(row, "shadow vlan required") or _get_cell(row, "shadow vlan required?"), True),
        "access_type": "EoFTTP",
        "bearer": bearer_b,
        "bandwidth": bandwidth_str,
        "contract_term": _months_to_term(term_months or "36"),
        "pay_upfront": _yes_no(_get_cell(row, "pay upfront") or _get_cell(row, "up-front charge") or _get_cell(row, "upfront charge") or _get_cell(row, "upfront cost (yes/no)") or _get_cell(row, "pay up-front circuit charge"), True),
        "fttp_aggregation": _yes_no(_get_cell(row, "fttp aggregation"), False),
        # RO2 diversity temporarily disabled in automation
        "ro2_diversity": False,
        # Demo 3/4: negotiated prices (discount = list - negotiated; list from page or optional columns)
        "negotiated_annual": _get_cell(row, "negotiated annual"),
        "negotiated_install": _get_cell(row, "negotiated install"),
        "list_annual": _get_cell(row, "list annual"),
        "list_install": _get_cell(row, "list install"),
    }
    billing = {
        "end_company": _cell(row, ["company name", "end company name"], billing_defaults, "end_company", "Test Co"),
        "first_name": _cell(row, ["contact first name", "first name"], billing_defaults, "first_name", "Test"),
        "surname": _cell(row, ["contact surname", "surname"], billing_defaults, "surname", "User"),
        "phone": _cell(row, ["phone", "mobile or landline phone number"], billing_defaults, "phone", "07123456789"),
        "email": _cell(row, ["email", "email address"], billing_defaults, "email", "test@example.com"),
        "rack_id": _cell(row, ["rack id"], billing_defaults, "rack_id", "1"),
        "vlan_id": _cell(row, ["vlan id"], billing_defaults, "vlan_id", "100"),
        "shadow_vlan_id": _cell(row, ["shadow vlan id"], billing_defaults, "shadow_vlan_id", "100"),
        "po_ref": _cell(row, ["po reference", "po ref"], billing_defaults, "po_ref", "TEST-PO-001"),
        "floor": _cell(row, ["floor"], billing_defaults, "floor", "001 - 1st Floor"),
        "room": _cell(row, ["room"], billing_defaults, "room", "ADMN - Admin Room"),
    }
    base["billing"] = billing
    # Site config
    ct = (_cell(row, ["connector type"], billing_defaults, "connector_type", "LC") or "LC").upper()
    ct = ct if ct in ("LC", "SC", "RJ45") else "LC"
    ps = (_cell(row, ["power supply", "power"], billing_defaults, "power_supply", "AC") or "AC").upper()
    ps = ps if ps in ("AC", "DC") else "AC"
    mt_raw = _cell(row, ["media type"], billing_defaults, "media_type", "LR") or "LR"
    mt = mt_raw.upper()
    if mt in ("LR", "LX", "SINGLE") or "single" in mt_raw.lower():
        mt = "LR"
    elif mt in ("SR", "SX", "MULTI") or "multi" in mt_raw.lower():
        mt = "SR"
    else:
        mt = "LR" if mt != "TX" else "TX"
    an_raw = _cell(row, ["access notice"], billing_defaults, "access_notice", "0-48 hours")
    access_notice = "Over 48 hours" if an_raw and "over" in an_raw.lower() else "0-48 hours"
    base["site_config"] = {
        "connector_type": ct,
        "power_supply": ps,
        "media_type": mt,
        "vlan_tagging": _yes_no(_cell(row, ["vlan tagging"], billing_defaults, "vlan_tagging"), False),
        "access_notice": access_notice,
        "auto_negotiation": _yes_no(_cell(row, ["auto negotiation"], billing_defaults, "auto_negotiation"), True),
        "hazards_on_site": _yes_no(_cell(row, ["hazards on site"], billing_defaults, "hazards_on_site"), True),
        "hazards_description": _cell(row, ["hazards description"], billing_defaults, "hazards_description", "Standard building hazards – site survey recommended prior to engineer visit.") or "Standard building hazards – site survey recommended prior to engineer visit.",
        "building_built_prior_2000": _yes_no(_cell(row, ["building built prior 2000"], billing_defaults, "building_built_prior_2000"), False),
        "asbestos_register": _yes_no(_cell(row, ["asbestos register"], billing_defaults, "asbestos_register"), False),
        "more_than_one_tenant": _yes_no(_cell(row, ["more than one tenant"], billing_defaults, "more_than_one_tenant"), False),
        "land_owner_permission_required": _yes_no(_cell(row, ["land owner permission required"], billing_defaults, "land_owner_permission_required"), False),
    }
    return base


def validate_csv(path: Path) -> tuple[list[dict], list[str]]:
    """
    Validate CSV and return (rows, errors).
    Supports two formats:
      - Simple: preset, postcode (rows have preset_path, preset_id, postcode)
      - Full: Product Type, B-End Postcode, Bearer, Bandwidth, etc. (rows have preset_dict, preset_id, postcode)
    """
    errors = []
    rows = []
    if not path.exists():
        return [], [f"File not found: {path}"]
    try:
        raw = path.read_text(encoding="utf-8-sig")
        # Detect delimiter: Excel/Numbers in some locales use semicolon
        first_line = raw.split("\n")[0] if raw else ""
        delimiter = ";" if ";" in first_line and first_line.count(";") > first_line.count(",") else ","
        reader = csv.DictReader(raw.splitlines(), delimiter=delimiter)
        raw_rows = list(reader)
        headers = [str(c).strip().lower().replace("\ufeff", "") for c in (reader.fieldnames or []) if c]

        if not headers:
            return [], ["CSV has no headers"]

        # Detect format: full has Product Type, B-End Postcode, Bearer/Bandwidth, Connector Type, etc.
        full_indicators = (
            "product type", "b-end postcode", "b_end postcode", "postcode",
            "bearer", "bandwidth (mbps)", "bandwidth", "b-end port (mbps)",
            "connector type", "company name", "end company name"
        )
        is_full = any(h in headers for h in full_indicators)

        if is_full:
            billing_defaults = {}
            for i, row in enumerate(raw_rows):
                preset_num = i + 1
                excel_row = i + 2
                row_label = f"Preset {preset_num} (Excel row {excel_row})"
                if i >= MAX_ROWS:
                    errors.append(f"Max {MAX_ROWS} rows allowed. {row_label} skipped.")
                    break
                b_end = _get_cell(row, "b-end postcode") or _get_cell(row, "b_end postcode") or _get_cell(row, "postcode")
                if not b_end and i > 0:
                    b_end = (billing_defaults.get("b_end_postcode") or "").strip()
                if not b_end:
                    errors.append(f"{row_label}: B-End Postcode is required")
                    continue
                if i == 0:
                    def _d(c): return _get_cell(row, c)
                    billing_defaults = {
                        "b_end_postcode": b_end,
                        "end_company": _d("company name") or _d("end company name"),
                        "first_name": _d("contact first name") or _d("first name"),
                        "surname": _d("contact surname") or _d("surname"),
                        "phone": _d("phone"),
                        "email": _d("email") or _d("email address"),
                        "rack_id": _d("rack id"),
                        "vlan_id": _d("vlan id"),
                        "shadow_vlan_id": _d("shadow vlan id"),
                        "po_ref": _d("po reference") or _d("po ref"),
                        "floor": _d("floor"),
                        "room": _d("room"),
                        "connector_type": _d("connector type"),
                        "power_supply": _d("power supply"),
                        "media_type": _d("media type"),
                        "vlan_tagging": _d("vlan tagging"),
                        "access_notice": _d("access notice"),
                        "auto_negotiation": _d("auto negotiation"),
                        "hazards_on_site": _d("hazards on site"),
                        "hazards_description": _d("hazards description"),
                        "building_built_prior_2000": _d("building built prior 2000"),
                        "asbestos_register": _d("asbestos register"),
                        "more_than_one_tenant": _d("more than one tenant"),
                        "land_owner_permission_required": _d("land owner permission required"),
                    }
                preset_dict = _build_preset_from_row(row, billing_defaults, i + 1)
                # Bearer/bandwidth constraint validation
                try:
                    from p2nni_constraints import validate_row
                    q = preset_dict.get("quote", {})
                    sc = preset_dict.get("site_config", {})
                    an = sc.get("auto_negotiation")
                    vt = sc.get("vlan_tagging")
                    errs = validate_row(
                        bearer=q.get("bearer"),
                        bandwidth=q.get("bandwidth"),
                        connector_type=sc.get("connector_type"),
                        media_type=sc.get("media_type"),
                        power_supply=sc.get("power_supply"),
                        auto_negotiation="Yes" if an else "No",
                        vlan_tagging="Yes" if vt else "No",
                    )
                    for e in errs:
                        errors.append(f"{row_label}: {e}")
                except ImportError:
                    pass
                postcode = _get_cell(row, "b-end postcode") or _get_cell(row, "b_end postcode")
                rows.append({"preset_dict": preset_dict, "preset_id": preset_dict["id"], "postcode": postcode or None})
        else:
            if "preset" not in headers:
                return [], ["CSV must have 'preset' and 'postcode' columns, or use full format (Product Type, B-End Postcode, etc.). See CSV_UPLOAD_FORMAT.md"]
            for i, row in enumerate(raw_rows, start=2):
                preset_num = i - 1
                row_label = f"Preset {preset_num} (Excel row {i})"
                if i - 1 > MAX_ROWS:
                    errors.append(f"Max {MAX_ROWS} rows allowed. {row_label} skipped.")
                    break
                preset_val = re.sub(r'^["\']|["\']$', "", _get_cell(row, "preset"))
                postcode_val = re.sub(r'^["\']|["\']$', "", _get_cell(row, "postcode"))
                if not preset_val:
                    errors.append(f"{row_label}: preset is empty")
                    continue
                resolved = resolve_preset(preset_val)
                if not resolved:
                    errors.append(f"{row_label}: preset '{preset_val}' not found")
                    continue
                rows.append({"preset_path": resolved, "preset_id": resolved.stem, "postcode": postcode_val or None})
    except Exception as e:
        return [], [f"Could not read CSV: {e}"]
    return rows, errors


def _run_preset_in_process_impl(
    preset_path: Path,
    postcode_override: str | None,
    suppress_output: bool,
    headless: bool,
):
    """Runs the real run_preset in-process; returns order/quote urls so Demo 2 can open orders."""
    import contextlib
    import io
    import traceback
    from run_preset import run_preset as _run_preset_impl

    start = time.perf_counter()
    order_id = quotation_id = line_id = tcv_total = start_supplier = ""
    install_price = annual_rental = ftpp_aggregation = add_on = ""
    order_number = quote_number = order_url = quote_url = ""
    out = io.StringIO() if suppress_output else None
    try:
        with (
            contextlib.redirect_stdout(out or sys.stdout),
            contextlib.redirect_stderr(out or sys.stderr),
        ):
            ret = _run_preset_impl(preset_path, postcode_override, headless=headless, pause=False)
        if isinstance(ret, tuple) and len(ret) >= 3:
            order_id = (ret[0] or "") if ret[0] is not None else ""
            quotation_id = (ret[1] or "") if len(ret) > 1 and ret[1] is not None else ""
            line_id = (ret[2] or "") if len(ret) > 2 and ret[2] is not None else ""
            tcv_total = (ret[3] or "") if len(ret) > 3 and ret[3] is not None else ""
            start_supplier = (ret[4] or "") if len(ret) > 4 and ret[4] is not None else ""
            install_price = (ret[5] or "") if len(ret) > 5 and ret[5] is not None else ""
            annual_rental = (ret[6] or "") if len(ret) > 6 and ret[6] is not None else ""
            ftpp_aggregation = (ret[7] or "") if len(ret) > 7 and ret[7] is not None else ""
            add_on = (ret[8] or "") if len(ret) > 8 and ret[8] is not None else ""
            if len(ret) > 9:
                order_number = (ret[9] or "") if ret[9] is not None else ""
            if len(ret) > 10:
                quote_number = (ret[10] or "") if ret[10] is not None else ""
            if len(ret) > 11:
                order_url = (ret[11] or "") if ret[11] is not None else ""
            if len(ret) > 12:
                quote_url = (ret[12] or "") if ret[12] is not None else ""
        elif ret is not None:
            order_id = str(ret)
        return (0, time.perf_counter() - start, "", order_id, quotation_id, line_id, tcv_total, start_supplier, install_price, annual_rental, ftpp_aggregation, add_on, order_number, quote_number, order_url, quote_url)
    except Exception:
        duration = time.perf_counter() - start
        err = traceback.format_exc()[-800:].strip()
        return (1, duration, err, "", "", "", "", "", "", "", "", "", "", "", "", "")


def _run_preset_demo34_impl(
    preset_path: Path,
    postcode_override: str | None,
    capture_basket_id: bool,
    suppress_output: bool,
) -> tuple[int, float, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str]:
    """Runs Demo 3/4 flow; returns 17 elements (same 16 as run_preset + basket_id)."""
    import contextlib
    import io
    import os
    import traceback
    from run_preset import run_preset_demo3_demo4

    start = time.perf_counter()
    out = io.StringIO() if suppress_output else None
    try:
        with (
            contextlib.redirect_stdout(out or sys.stdout),
            contextlib.redirect_stderr(out or sys.stderr),
        ):
            # Allow overriding Demo3/4 headless mode via env var so you can watch the browser.
            # P2NNI_DEMO34_HEADLESS: "0" = headful (visible), anything else/default = headless.
            headless_env = os.getenv("P2NNI_DEMO34_HEADLESS", "1")
            headless_flag = False if headless_env == "0" else True
            ret = run_preset_demo3_demo4(
                preset_path,
                postcode_override,
                capture_basket_id=capture_basket_id,
                headless=headless_flag,
            )
        # ret: (order_id, quotation_num, line_id, tcv_total, start_supplier, install_price, annual_rental,
        #       ftpp_aggregation, add_on, order_number, quote_number, order_url, quote_url, basket_id)
        order_id = (ret[0] or "") if ret else ""
        quotation_id = (ret[1] or "") if len(ret) > 1 else ""
        line_id = (ret[2] or "") if len(ret) > 2 else ""
        tcv_total = (ret[3] or "") if len(ret) > 3 else ""
        start_supplier = (ret[4] or "") if len(ret) > 4 else ""
        install_price = (ret[5] or "") if len(ret) > 5 else ""
        annual_rental = (ret[6] or "") if len(ret) > 6 else ""
        ftpp_aggregation = (ret[7] or "") if len(ret) > 7 else ""
        add_on = (ret[8] or "") if len(ret) > 8 else ""
        order_number = (ret[9] or "") if len(ret) > 9 else ""
        quote_number = (ret[10] or "") if len(ret) > 10 else ""
        order_url = (ret[11] or "") if len(ret) > 11 else ""
        quote_url = (ret[12] or "") if len(ret) > 12 else ""
        basket_id = (ret[13] or "") if len(ret) > 13 else ""
        return (
            0, time.perf_counter() - start, "",
            order_id, quotation_id, line_id, tcv_total, start_supplier,
            install_price, annual_rental, ftpp_aggregation, add_on,
            order_number, quote_number, order_url, quote_url, basket_id,
        )
    except Exception:
        duration = time.perf_counter() - start
        err = traceback.format_exc()[-800:].strip()
        return (1, duration, err, "", "", "", "", "", "", "", "", "", "", "", "", "", "")


def run_preset(preset_path: Path, postcode_override: str | None, headless: bool = True, suppress_output: bool = True) -> tuple[int, float, str, str, str, str, str, str, str, str, str, str, str, str, str, str]:
    """Returns (exit_code, duration_sec, stderr_snippet, order_id, quotation_id, line_id, tcv_total, start_supplier, install_price, annual_rental, ftpp_aggregation, add_on, order_number, quote_number, order_url, quote_url)."""
    try:
        from run_preset import run_preset as _run_preset_impl
    except ImportError:
        _run_preset_impl = None

    # Always prefer in-process execution so we can return scraped fields like `order_url`.
    # This is critical for Demo 2 when customer run is headful ("Show browser").
    if _run_preset_impl:
        start = time.perf_counter()
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(
                _run_preset_in_process_impl,
                preset_path,
                postcode_override,
                suppress_output,
                headless,
            )
            try:
                return future.result(timeout=PRESET_TIMEOUT_SEC)
            except FuturesTimeoutError:
                duration = time.perf_counter() - start
                timeout_msg = (
                    f"Preset timed out after {PRESET_TIMEOUT_SEC // 60} minutes. "
                    "Portal may be slow or the UI may have changed. "
                    "Run from terminal: python run_csv_regression.py your.csv --verbose to see where it stops."
                )
                return (1, duration, timeout_msg, "", "", "", "", "", "", "", "", "", "", "", "", "")

    cmd = ["python3", str(RUN_PRESET_SCRIPT), str(preset_path)]
    if postcode_override:
        cmd.extend(["--postcode", postcode_override.strip()])
    if headless:
        cmd.append("--headless")
    start = time.perf_counter()
    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        timeout=600,
    )
    duration = time.perf_counter() - start
    err = ""
    if result.returncode != 0 and result.stderr:
        err = (result.stderr[-800:] if len(result.stderr) > 800 else result.stderr).strip()
    return result.returncode, duration, err, "", "", "", "", "", "", "", "", "", "", "", "", ""


def prompt_for_csv() -> Path | None:
    """User-friendly prompt for CSV path."""
    print("\n" + "=" * 60)
    print("  CSV Regression Upload")
    print("=" * 60)
    print("\n  Enter the path to your CSV file.")
    print("  (You can drag and drop the file into this window)")
    print("\n  Example: /Users/you/Downloads/my_regression.csv")
    print("  Example: example_regression_upload.csv")
    print("\n  Format: See CSV_UPLOAD_FORMAT.md")
    print("  Sample:  example_regression_upload.csv")
    print()
    raw = input("  Path to CSV: ").strip()
    raw = re.sub(r'^["\']|["\']$', "", raw)
    if not raw:
        return None
    p = Path(raw)
    if not p.is_absolute():
        p = (Path.cwd() / p).resolve()
    return p if p.exists() else None


# Power BI / SharePoint bridge: base headers for live data.
SUMMARY_CSV_BASE_HEADERS = [
    "Preset ID",
    "StartPostCode",
    "EndPostCode",
    "CompanyName",
    "Usage",
    "Start Supplier",
    "End Supplier",
    "StartPort Speed",
    "EndPort Speed",
    "PathSpeed",
    "TermLength",
    "InstallPrice",
    "AnnualRental",
    "FTTP Aggregation",
    "Quote",
    "Order",
    "Result",
    "Duration (seconds)",
    "Date Completed",
    "Error",
]


def _summary_headers_for_mode(mode: str) -> list[str]:
    m = (mode or "").strip().lower()
    headers = list(SUMMARY_CSV_BASE_HEADERS)
    if m in ("demo3", "demo4"):
        insert_at = headers.index("Quote")
        headers[insert_at:insert_at] = [
            "NegotiatedAnnual",
            "NegotiatedInstall",
            "ListAnnual",
            "ListInstall",
        ]
    if m in ("demo2", "demo4"):
        insert_at = headers.index("Result")
        headers.insert(insert_at, "BasketId")
    return headers


def _format_tcv(raw: str) -> str:
    """Format TCV for CSV: e.g. '16216.20' -> '£16,216.20'."""
    s = (raw or "").strip().replace(",", "")
    if not s:
        return ""
    try:
        val = float(s)
        return f"£{val:,.2f}"
    except ValueError:
        return raw


def _format_currency(raw: str) -> str:
    """Format currency for CSV: e.g. '4072.00' -> '£4,072.00'."""
    try:
        s = (raw or "").strip().replace(",", "")
        if not s:
            return ""
        val = float(s)
        return f"£{val:,.2f}"
    except (ValueError, TypeError):
        return str(raw or "")


def _currency_numeric(raw: str) -> str:
    """Currency as plain number string for Power BI (no £, no commas).

    Examples:
      - '0.00' -> '0'
      - '4072.00' -> '4072'
      - '4054.05' -> '4054.05'
    """
    try:
        s = (raw or "").strip().replace(",", "").lstrip("£")
        if not s:
            return ""
        val = float(s)
        # Drop .00 for whole numbers, keep 2dp otherwise
        if val.is_integer():
            return str(int(val))
        return f"{val:.2f}"
    except (ValueError, TypeError):
        return (raw or "").strip()


def _term_length_for_powerbi(raw: str) -> str:
    """Normalise contract_term to numeric months for Power BI (e.g. '1 year' -> '12', '36 months' -> '36')."""
    s = (raw or "").strip()
    if not s:
        return ""
    s_lower = s.lower()
    if "year" in s_lower:
        m = re.search(r"(\d+)", s)
        if m:
            return str(int(m.group(1)) * 12)
    if "month" in s_lower or s.isdigit():
        m = re.search(r"\d+", s)
        if m:
            return m.group(0)
    return s


def _excel_escape(s: str) -> str:
    """Escape double quotes for use inside an Excel formula string."""
    return (s or "").replace('"', '""')


def _extract_basket_id_from_page(page) -> str:
    """Extract Basket Id from the current order page, using the same style as Install/Annual scraping."""
    try:
        body = page.evaluate("() => document.body ? (document.body.innerText || '') : ''") or ""
        # Example snippet (from you): "Basket Id: 220065"
        # Be forgiving on spacing and case: "Basket Id", "Basket ID", etc.
        m = re.search(r"Basket\s*I[Dd]:\s*([0-9]+)", body, re.IGNORECASE)
        if not m:
            # Fallback without colon, in case the UI ever renders "Basket Id 220075"
            m = re.search(r"Basket\s*I[Dd]\s*([0-9]+)", body, re.IGNORECASE)
        if m:
            return m.group(1)
    except Exception:
        return ""
    return ""


def _write_summary_xlsx(results: list, summary_path_xlsx: Path, date_str: str, headers: list[str]) -> None:
    """Write summary to Excel with Order and Quote as clickable HYPERLINK formulas."""
    if not OPENPYXL_AVAILABLE:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    # Track max text length per column to size widths to content (esp. Order/Quote)
    max_len: dict[str, int] = {h: len(h) for h in headers}
    order_col = headers.index("Order") + 1 if "Order" in headers else 0
    quote_col = headers.index("Quote") + 1 if "Quote" in headers else 0
    for row_idx, r in enumerate(results, start=2):
        row_data = {
            "Preset ID": r.preset_id,
            "StartPostCode": r.postcode or "",
            "EndPostCode": r.postcode or "",
            "CompanyName": getattr(r, "company_name", "") or "",
            "Usage": "Point-to-Network-to-Network Interface",
            "Start Supplier": "",
            "End Supplier": getattr(r, "start_supplier", "") or "BT Openreach",
            "StartPort Speed": getattr(r, "start_port_speed", "") or "",
            "EndPort Speed": getattr(r, "end_port_speed", "") or "",
            "PathSpeed": getattr(r, "path_speed", "") or "",
            "TermLength": _term_length_for_powerbi(getattr(r, "term_length", "") or ""),
            "InstallPrice": _format_currency(getattr(r, "install_price", "") or ""),
            "AnnualRental": _format_currency(getattr(r, "annual_rental", "") or ""),
            "FTTP Aggregation": _format_currency(getattr(r, "ftpp_aggregation", "") or ""),
            "BasketId": getattr(r, "basket_id", "") or "",
            "NegotiatedAnnual": _format_currency(getattr(r, "negotiated_annual", "") or ""),
            "NegotiatedInstall": _format_currency(getattr(r, "negotiated_install", "") or ""),
            "ListAnnual": _format_currency(getattr(r, "list_annual", "") or ""),
            "ListInstall": _format_currency(getattr(r, "list_install", "") or ""),
            "Order": getattr(r, "order_number", "") or "",
            "Quote": getattr(r, "quote_number", "") or "",
            "Result": r.result,
            "Duration (seconds)": round(r.duration_sec, 1),
            "Date Completed": getattr(r, "completed_at", date_str) or date_str,
            "Error": (getattr(r, "error_detail", "") or "")[:2000],
        }
        for col_idx, header in enumerate(headers, start=1):
            val = row_data.get(header, "")
            # For width, use display text length; for hyperlinks use ID text, not full formula
            if header == "Order":
                display_text = getattr(r, "order_number", "") or ""
            elif header == "Quote":
                display_text = getattr(r, "quote_number", "") or ""
            else:
                display_text = str(val or "")
            max_len[header] = max(max_len.get(header, 0), len(display_text))
            if col_idx == order_col and isinstance(val, str) and val:
                order_url = (getattr(r, "order_url", "") or "").strip()
                if order_url:
                    formula = f'=HYPERLINK("{_excel_escape(order_url)}", "{_excel_escape(val)}")'
                    ws.cell(row=row_idx, column=col_idx, value=formula)
                else:
                    ws.cell(row=row_idx, column=col_idx, value=val)
            elif col_idx == quote_col and isinstance(val, str) and val:
                quote_url = (getattr(r, "quote_url", "") or "").strip()
                if quote_url:
                    formula = f'=HYPERLINK("{_excel_escape(quote_url)}", "{_excel_escape(val)}")'
                    ws.cell(row=row_idx, column=col_idx, value=formula)
                else:
                    ws.cell(row=row_idx, column=col_idx, value=val)
            else:
                ws.cell(row=row_idx, column=col_idx, value=val)
    # Apply column widths based on max text length (cap to avoid huge "Error" column)
    for col_idx, header in enumerate(headers, start=1):
        from openpyxl.utils import get_column_letter
        width = min(max_len.get(header, len(header)) + 2, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(summary_path_xlsx)


def print_summary(
    results: list[RowResult],
    total_duration: float,
    summary_path: Path | None,
    summary_path_powerbi: Path | None = None,
    run_id: str | None = None,
    mode: str = "demo1",
):
    """Print and optionally save summary. run_id: same for all rows in this run (for Power BI)."""
    passed = sum(1 for r in results if r.exit_code == 0)
    failed = len(results) - passed
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    run_id = run_id or date_str
    headers = _summary_headers_for_mode(mode)

    print("\n" + "=" * 60)
    print("  REGRESSION SUMMARY")
    print("=" * 60)
    print(f"\n  Status:   {'All passed' if failed == 0 else f'{failed} failed'}")
    print(f"  Passed:   {passed}/{len(results)}")
    print(f"  Failed:   {failed}/{len(results)}")
    print(f"  Duration: {total_duration/60:.1f} min")
    print(f"  Date:     {date_str}")
    print("\n  Results:")
    for r in results:
        icon = "  ✓" if r.exit_code == 0 else "  ✗"
        print(f"    {icon} {r.preset_id}  ({r.duration_sec:.0f}s)")
    print("=" * 60)

    if summary_path:
        RESULTS_DIR.mkdir(exist_ok=True)
        # Power BI variant: same headers but numeric currency fields
        summary_powerbi = summary_path_powerbi or summary_path.with_name(summary_path.stem + "_powerbi.csv")
        with open(summary_path, "w", newline="", encoding="utf-8") as f_pretty, \
             open(summary_powerbi, "w", newline="", encoding="utf-8") as f_pbi:
            w_pretty = csv.DictWriter(f_pretty, fieldnames=headers, extrasaction="ignore")
            w_pbi = csv.DictWriter(f_pbi, fieldnames=headers, extrasaction="ignore")
            w_pretty.writeheader()
            w_pbi.writeheader()
            for r in results:
                # Build Excel-friendly hyperlink formulas for CSV so that Order/Quote are clickable
                order_number = getattr(r, "order_number", "") or ""
                quote_number = getattr(r, "quote_number", "") or ""
                order_url = (getattr(r, "order_url", "") or "").strip()
                quote_url = (getattr(r, "quote_url", "") or "").strip()
                order_cell = order_number
                quote_cell = quote_number
                if order_number and order_url:
                    order_cell = f'=HYPERLINK("{_excel_escape(order_url)}", "{_excel_escape(order_number)}")'
                if quote_number and quote_url:
                    quote_cell = f'=HYPERLINK("{_excel_escape(quote_url)}", "{_excel_escape(quote_number)}")'

                install_raw = getattr(r, "install_price", "") or ""
                annual_raw = getattr(r, "annual_rental", "") or ""
                ftpp_raw = getattr(r, "ftpp_aggregation", "") or ""

                base_row = {
                    "Preset ID": r.preset_id,
                    "StartPostCode": r.postcode or "",
                    "EndPostCode": r.postcode or "",
                    "CompanyName": getattr(r, "company_name", "") or "",
                    "Usage": "Point-to-Network-to-Network Interface",
                    "Start Supplier": "",
                    "End Supplier": getattr(r, "start_supplier", "") or "BT Openreach",
                    "StartPort Speed": getattr(r, "start_port_speed", "") or "",
                    "EndPort Speed": getattr(r, "end_port_speed", "") or "",
                    "PathSpeed": getattr(r, "path_speed", "") or "",
                    "TermLength": _term_length_for_powerbi(getattr(r, "term_length", "") or ""),
                    "InstallPrice": _format_currency(install_raw),
                    "AnnualRental": _format_currency(annual_raw),
                    "FTTP Aggregation": _format_currency(ftpp_raw),
                    "BasketId": getattr(r, "basket_id", "") or "",
                    "NegotiatedAnnual": _format_currency(getattr(r, "negotiated_annual", "") or ""),
                    "NegotiatedInstall": _format_currency(getattr(r, "negotiated_install", "") or ""),
                    "ListAnnual": _format_currency(getattr(r, "list_annual", "") or ""),
                    "ListInstall": _format_currency(getattr(r, "list_install", "") or ""),
                    "Order": order_cell,
                    "Quote": quote_cell,
                    "Result": r.result,
                    "Duration (seconds)": round(r.duration_sec, 1),
                    "Date Completed": getattr(r, "completed_at", date_str) or date_str,
                    "Error": (r.error_detail or "")[:2000],
                }

                w_pretty.writerow(base_row)

                base_row_pbi = dict(base_row)
                base_row_pbi["BasketId"] = getattr(r, "basket_id", "") or ""
                # For Power BI: plain IDs in Order/Quote (no HYPERLINK formulas)
                base_row_pbi["Order"] = order_number
                base_row_pbi["Quote"] = quote_number
                base_row_pbi["InstallPrice"] = _currency_numeric(install_raw)
                base_row_pbi["AnnualRental"] = _currency_numeric(annual_raw)
                base_row_pbi["FTTP Aggregation"] = _currency_numeric(ftpp_raw)
                base_row_pbi["NegotiatedAnnual"] = _currency_numeric(getattr(r, "negotiated_annual", "") or "")
                base_row_pbi["NegotiatedInstall"] = _currency_numeric(getattr(r, "negotiated_install", "") or "")
                base_row_pbi["ListAnnual"] = _currency_numeric(getattr(r, "list_annual", "") or "")
                base_row_pbi["ListInstall"] = _currency_numeric(getattr(r, "list_install", "") or "")
                w_pbi.writerow(base_row_pbi)

        print(f"\n  Summary saved to: {summary_path}")
        print(f"  Power BI summary (numeric currency) saved to: {summary_powerbi}")
        # Also write Excel with clickable Order/Quote hyperlinks
        summary_xlsx = summary_path.with_suffix(".xlsx")
        try:
            _write_summary_xlsx(results, summary_xlsx, date_str, headers)
            print(f"  Excel summary (clickable Order/Quote links): {summary_xlsx}\n")
        except Exception as e:
            print(f"  (Excel summary skipped: {e})\n")


def fill_basket_ids_with_internal_user(results: list[RowResult], mode: str) -> None:
    """For Demo 2: log in as internal NEOS user and capture BasketId for each successful order."""
    if mode != "demo2":
        return
    if not results:
        return
    creds = get_neos_internal_creds()
    if not creds:
        print("ℹ️ Demo 2 requested but neos_internal credentials are missing in config.json — skipping BasketId collection.")
        return
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError, expect  # type: ignore
    except Exception:
        print("ℹ️ Demo 2 requested but Playwright is not available in this environment — skipping BasketId collection.")
        return

    base_url = get_portal_base_url().rstrip("/")
    print("\n" + "=" * 60)
    print("  DEMO 2: Internal login to capture BasketId")
    print("=" * 60)

    # For Demo 2 we keep the internal NEOS browser headless by default.
    # The main customer run visibility is controlled elsewhere (P2NNI_DEMO2_HEADLESS).
    # If you specifically want to watch the internal steps, set:
    #   P2NNI_DEMO2_INTERNAL_HEADLESS=0
    internal_headless = os.getenv("P2NNI_DEMO2_INTERNAL_HEADLESS", "1") != "0"
    slow_mo = 0 if internal_headless else 200
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=internal_headless, slow_mo=slow_mo)
        context = browser.new_context()
        page = context.new_page()
        page.set_default_timeout(10000)
        print(f"  Internal NEOS browser: headless={internal_headless} (slow_mo={slow_mo})")

        def _login_internal():
            print("🔐 Logging in as NEOS internal user…")
            page.goto(f"{base_url}/login", wait_until="domcontentloaded", timeout=60000)
            # Try common selectors for email/password + sign-in button.
            # We keep this very defensive because the login screen is Cognito-hosted.

            candidates_email = [
                page.get_by_label("Email"),
                page.get_by_label(re.compile("email", re.IGNORECASE)),
                page.get_by_placeholder(re.compile("email", re.IGNORECASE)),
                page.get_by_label("Username"),
                page.get_by_label(re.compile("username", re.IGNORECASE)),
                page.get_by_placeholder(re.compile("username", re.IGNORECASE)),
                page.locator("input[type='email']"),
                page.locator("input[name*='email' i]"),
                page.locator("input[id*='email' i]"),
                page.locator("input[name*='username' i]"),
                page.locator("input[id*='username' i]"),
            ]
            email_box = None
            for c in candidates_email:
                try:
                    if c.count() > 0 and c.first.is_visible(timeout=2000):
                        email_box = c.first
                        break
                except Exception:
                    continue

            candidates_pass = [
                page.get_by_label("Password"),
                page.get_by_label(re.compile("password", re.IGNORECASE)),
                page.get_by_placeholder(re.compile("password", re.IGNORECASE)),
                page.locator("input[type='password']"),
                page.locator("input[name*='password' i]"),
                page.locator("input[id*='password' i]"),
            ]
            pass_box = None
            for c in candidates_pass:
                try:
                    if c.count() > 0 and c.first.is_visible(timeout=2000):
                        pass_box = c.first
                        break
                except Exception:
                    continue
            if not email_box or not pass_box:
                print("⚠️ Could not find email/password fields on login page — BasketId capture skipped.")
                return False
            email_box.fill(creds["email"])
            pass_box.fill(creds["password"])
            # Look for a primary login/sign-in button
            btn = None
            for b in [
                page.get_by_role("button", name=re.compile("sign in", re.IGNORECASE)),
                page.get_by_role("button", name=re.compile("log in|login", re.IGNORECASE)),
                page.get_by_role("button"),
            ]:
                try:
                    if b.count() > 0 and b.first.is_visible(timeout=5000):
                        btn = b.first
                        break
                except Exception:
                    continue
            if not btn:
                print("⚠️ Could not find Sign in button — BasketId capture skipped.")
                return False
            btn.click()
            try:
                page.wait_for_url(re.compile(r"/orders|/welcome|/dashboard"), timeout=60000)
            except PWTimeoutError:
                pass
            print(f"✅ Logged in as internal user. URL: {page.url}")
            return True

        if not _login_internal():
            context.close()
            browser.close()
            return

        # Process each successful order (needs order_url so we can open the order directly)
        successful = [r for r in results if r.result == "Pass" and r.order_url]
        total = len(successful)
        print(f"  ✅ Internal candidates with order_url: {total}/{len(results)}")
        for idx, r in enumerate(successful, 1):
            print(f"\n[Demo 2] ({idx}/{total}) Processing order for preset {r.preset_id}…")
            # Open the order directly via its URL — this matches how you navigate back after placing.
            try:
                print(f"  🔗 Opening order directly via URL: {r.order_url}")
                page.goto(r.order_url, wait_until="domcontentloaded", timeout=60000)
            except Exception as e:
                print(f"  ⚠️ Could not open order URL {r.order_url}: {e}")
                continue

            # If BasketId already present, capture and continue
            try:
                initial_basket = _extract_basket_id_from_page(page)
            except Exception:
                initial_basket = ""
            if initial_basket:
                r.basket_id = initial_basket
                print(f"  ✅ BasketId already present: {r.basket_id}")
                continue

            # FULLY AUTOMATED: click "Place order" and "OK" as NEOS internal user.
            did_click_place = False
            try:
                # Scroll near bottom where the Review & place order buttons live
                try:
                    page.evaluate("window.scrollTo(0, document.body.scrollHeight);")
                    page.wait_for_timeout(800)
                except Exception:
                    pass

                # Primary locator: button.accept-btn (from DevTools)
                btn = page.locator("button.accept-btn").first
                try:
                    # Wait up to 60s for the button to appear at all
                    btn.wait_for(state="visible", timeout=60000)
                except Exception:
                    # Fallback: any button whose accessible name contains "Place order"
                    btn = page.get_by_role("button", name=re.compile(r"Place order", re.IGNORECASE)).first
                    btn.wait_for(state="visible", timeout=60000)

                print("  🟧 Clicking 'Place order' as internal user…")
                btn.scroll_into_view_if_needed(timeout=5000)
                try:
                    expect(btn).to_be_enabled(timeout=20000)
                except Exception:
                    print("  ℹ️ 'Place order' button may not report enabled; click will be attempted anyway.")
                btn.click(timeout=30000)
                did_click_place = True

                # Wait for "Order Accepted" dialog and click OK
                try:
                    page.get_by_text(re.compile(r"Order Accepted", re.IGNORECASE)).first.wait_for(
                        state="visible", timeout=30000
                    )
                except Exception:
                    print("  ℹ️ 'Order Accepted' heading not detected; continuing to look for OK button.")
                try:
                    ok_btn = page.get_by_role("button", name=re.compile(r"^OK$", re.IGNORECASE))
                    if ok_btn.count() > 0:
                        print("  🟧 Clicking 'OK' on Order Accepted dialog…")
                        ok = ok_btn.first
                        ok.scroll_into_view_if_needed(timeout=5000)
                        ok.click(timeout=30000)
                        print("  ✅ Order Accepted dialog confirmed.")
                    else:
                        print("  ℹ️ OK button not found on Order Accepted dialog; continuing anyway.")
                except Exception as e:
                    print(f"  ℹ️ Error while handling 'OK' button: {e}; continuing to poll for BasketId.")

                # After OK we are typically redirected away; go back to the order URL we care about.
                try:
                    page.goto(r.order_url, wait_until="domcontentloaded", timeout=60000)
                except Exception as e:
                    print(f"  ⚠️ Could not return to order URL after placing: {e}")

            except Exception as e:
                print(f"  ⚠️ Error while trying to click 'Place order' / 'OK': {e}")
                did_click_place = False

            # Poll for Basket Id: faster interval, similar total window (Demo 2 & 4).
            print("  ⏳ Waiting for Basket Id to appear (refreshing every ~3.2s, up to ~80s)…")
            basket_id = ""
            refresh_interval_ms = 3200
            max_attempts = 25  # ~80s total
            for attempt in range(1, max_attempts + 1):
                try:
                    if attempt == 1 or attempt % 3 == 0:
                        print(f"    ⏳ BasketId not ready yet (attempt {attempt}/{max_attempts})…")
                    page.goto(r.order_url, wait_until="domcontentloaded", timeout=60000)
                    page.wait_for_timeout(refresh_interval_ms)  # 5s between refreshes
                    basket_id = _extract_basket_id_from_page(page)
                    if basket_id:
                        break
                except Exception as e:
                    print(f"    ℹ️ Error while polling for Basket Id on attempt {attempt}: {e}")
                    continue
            if basket_id:
                r.basket_id = basket_id
                print(f"  ✅ BasketId captured: {basket_id}")
            else:
                print("  ⚠️ Basket Id not found within timeout; leaving BasketId blank for this row.")

        context.close()
        browser.close()

def run_csv_regression(
    csv_path: Path,
    *,
    quiet: bool = True,
    verbose: bool = False,
    progress_callback=None,
    progress_dict: dict | None = None,
    mode: str = "demo1",
) -> tuple[int, list[RowResult], Path | None, Path | None]:
    """
    Run regression from CSV. Returns (exit_code, results, summary_path).
    quiet=True: headless, minimal output, summary at end
    verbose=True: show per-preset progress
    """
    rows, errors = validate_csv(csv_path)
    if errors:
        for e in errors:
            print(f"  ⚠️ {e}")
        return 1, [], None, None
    if not rows:
        print("  ⚠️ No valid rows in CSV.")
        return 1, [], None, None

    # Confirm
    if verbose:
        print(f"\n  Found {len(rows)} preset(s) to run.")
        for i, r in enumerate(rows, 1):
            print(f"    {i}. {r['preset_id']}")
        print()
    else:
        print(f"\n  Found {len(rows)} preset(s).")

    if not RUN_PRESET_SCRIPT.exists():
        print(f"  ❌ Missing {RUN_PRESET_SCRIPT}")
        return 1, [], None, None

    if quiet:
        print("\n  Processing your CSV... (runs in background, no browser window)")
        print("  Please wait...\n")

    # Browser visibility defaults:
    #   - headless unless you explicitly set a demo-specific env var to "0".
    # Demo 1: allow visible browser via P2NNI_DEMO1_HEADLESS=0 (watch where time is spent)
    headless = True  # default: no browser popup
    if mode == "demo1":
        headless = os.getenv("P2NNI_DEMO1_HEADLESS", "1") != "0"
    if mode == "demo2":
        headless = os.getenv("P2NNI_DEMO2_HEADLESS", "1") != "0"
    if not headless and quiet:
        if mode == "demo1":
            print("  (Browser visible: P2NNI_DEMO1_HEADLESS=0)\n")
        elif mode == "demo2":
            print("  (Browser visible: P2NNI_DEMO2_HEADLESS=0)\n")
    suppress_output = quiet  # When quiet, hide preset script output; when verbose, show it
    results: list[RowResult] = []
    start_total = time.perf_counter()
    run_id_readable = datetime.now().strftime("%Y-%m-%d %H:%M")
    summary_path = RESULTS_DIR / "P2NNI_regression_summary.csv"
    summary_path_powerbi = RESULTS_DIR / "P2NNI_regression_summary_powerbi.csv"

    temp_paths: list[Path] = []
    total = len(rows)
    for i, row in enumerate(rows, 1):
        # Update progress so UI shows "Running preset X/Y (preset_id) — N min" during long runs
        if progress_dict is not None:
            progress_dict.update({
                "current_preset": row.get("preset_id", f"row_{i}"),
                "current_index": i,
                "current_started_at": time.time(),
            })
        if verbose:
            print(f"\n  [{i}/{len(rows)}] Running: {row['preset_id']}")
        elif quiet:
            print(f"  Running preset {i}/{len(rows)}...", end=" ", flush=True)
        preset_path: Path
        if "preset_dict" in row:
            tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False, encoding="utf-8")
            json.dump(row["preset_dict"], tmp, indent=0)
            tmp.close()
            preset_path = Path(tmp.name)
            temp_paths.append(preset_path)
        else:
            preset_path = row["preset_path"]
        use_demo34 = mode in ("demo3", "demo4")
        try:
            if use_demo34:
                with ThreadPoolExecutor(max_workers=1) as executor:
                    future = executor.submit(
                        _run_preset_demo34_impl,
                        preset_path,
                        row["postcode"],
                        capture_basket_id=(mode == "demo4"),
                        suppress_output=suppress_output,
                    )
                    ret = future.result(timeout=PRESET_TIMEOUT_SEC)
            else:
                ret = run_preset(preset_path, row["postcode"], headless=headless, suppress_output=suppress_output)
        except Exception as e:
            ret = (1, 0.0, str(e)[:500], "", "", "", "", "", "", "", "", "", "", "", "", "")
            if use_demo34:
                ret = ret + ("",)  # 17th element basket_id
        code = ret[0] if ret else 1
        duration = ret[1] if len(ret) > 1 else 0.0
        stderr_snippet = (ret[2] or "") if len(ret) > 2 else ""
        order_id = (ret[3] or "") if len(ret) > 3 else ""
        quotation_id = (ret[4] or "") if len(ret) > 4 else ""
        line_id = (ret[5] or "") if len(ret) > 5 else ""
        tcv_total = (ret[6] or "") if len(ret) > 6 else ""
        start_supplier = (ret[7] or "") if len(ret) > 7 else ""
        install_price = (ret[8] or "") if len(ret) > 8 else ""
        annual_rental = (ret[9] or "") if len(ret) > 9 else ""
        ftpp_aggregation = (ret[10] or "") if len(ret) > 10 else ""
        add_on = (ret[11] or "") if len(ret) > 11 else ""
        order_number = (ret[12] or "") if len(ret) > 12 else ""
        quote_number = (ret[13] or "") if len(ret) > 13 else ""
        order_url = (ret[14] or "") if len(ret) > 14 else ""
        quote_url = (ret[15] or "") if len(ret) > 15 else ""
        basket_id_from_run = (ret[16] or "") if len(ret) > 16 else ""
        # Retry once on bearer selection failure or Publish timeout (10 Gbps first row often fails due to cold-start)
        retry_triggers = ("Could not select bearer", "Publish button did not appear")
        if code != 0 and stderr_snippet and any(t in stderr_snippet for t in retry_triggers) and not use_demo34:
            if verbose:
                print(f"  Retrying {row['preset_id']} (bearer selection flake)...")
            time.sleep(2)
            try:
                ret = run_preset(preset_path, row["postcode"], headless=headless, suppress_output=suppress_output)
            except Exception as e:
                ret = (1, 0.0, str(e)[:500], "", "", "", "", "", "", "", "", "", "", "", "", "")
            code = ret[0] if ret else 1
            duration = ret[1] if len(ret) > 1 else 0.0
            stderr_snippet = (ret[2] or "") if len(ret) > 2 else ""
            order_id = (ret[3] or "") if len(ret) > 3 else ""
            quotation_id = (ret[4] or "") if len(ret) > 4 else ""
            line_id = (ret[5] or "") if len(ret) > 5 else ""
            tcv_total = (ret[6] or "") if len(ret) > 6 else ""
            start_supplier = (ret[7] or "") if len(ret) > 7 else ""
            install_price = (ret[8] or "") if len(ret) > 8 else ""
            annual_rental = (ret[9] or "") if len(ret) > 9 else ""
            ftpp_aggregation = (ret[10] or "") if len(ret) > 10 else ""
            add_on = (ret[11] or "") if len(ret) > 11 else ""
            order_number = (ret[12] or "") if len(ret) > 12 else ""
            quote_number = (ret[13] or "") if len(ret) > 13 else ""
            order_url = (ret[14] or "") if len(ret) > 14 else ""
            quote_url = (ret[15] or "") if len(ret) > 15 else ""
            basket_id_from_run = ""
        pd = row.get("preset_dict") or {}
        q = pd.get("quote") or {}
        billing = pd.get("billing") or {}
        def _port_mbps(val):
            """Normalise bearer/bandwidth to Mbps string for Power BI.

            Examples:
              - '10 Gbps' -> '10000'
              - '1 Gbps'  -> '1000'
              - '100 Mbps'-> '100'
              - '500 Mbps'-> '500'
            """
            if val is None:
                return ""
            s = str(val).strip().replace(" ", "").lower()
            if not s:
                return ""
            if s in ("10g", "10gbps", "10000"):
                return "10000"
            if s in ("1g", "1gbps", "1000"):
                return "1000"
            if s in ("100m", "100mbps", "100"):
                return "100"
            if s.isdigit():
                return str(int(s))
            if "mbps" in s or s.endswith("m"):
                digits = "".join(c for c in s if c.isdigit())
                return digits or ""
            return ""

        def _term_months(ct):
            """Contract term to numeric months for Power BI TermLength (e.g. '3 years' -> '36')."""
            if ct is None:
                return ""
            s = str(ct).strip().lower()
            if not s:
                return ""
            if s in ("1 year", "12 months"):
                return "12"
            if s in ("2 years", "24 months"):
                return "24"
            if s in ("3 years", "36 months"):
                return "36"
            if s in ("4 years", "48 months"):
                return "48"
            if s in ("5 years", "60 months"):
                return "60"
            m = re.search(r"(\d+)", s)
            return m.group(1) if m else ""

        result = RowResult(
            row=i,
            preset_id=row["preset_id"],
            postcode=row["postcode"] or "",
            exit_code=code,
            duration_sec=duration,
            result="Pass" if code == 0 else "Fail",
            error_detail=stderr_snippet or "",
            order_id=order_id or "",
            quotation_id=quotation_id or "",
            line_id=line_id or "",
            tcv_total=tcv_total or "",
            start_supplier=start_supplier or "",
            install_price=install_price or "",
            annual_rental=annual_rental or "",
            ftpp_aggregation=ftpp_aggregation or "",
            add_on=add_on or "",
            order_number=order_number or "",
            quote_number=quote_number or "",
            order_url=order_url or "",
            quote_url=quote_url or "",
            company_name=(billing.get("end_company") or "").strip(),
            start_port_speed=_port_mbps(q.get("bearer")),
            end_port_speed=_port_mbps(q.get("bearer")),
            path_speed=_port_mbps(q.get("bandwidth")),
            term_length=_term_months(q.get("contract_term")),
            completed_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            basket_id=basket_id_from_run or "",
            negotiated_annual=str(q.get("negotiated_annual") or "").strip(),
            negotiated_install=str(q.get("negotiated_install") or "").strip(),
            list_annual=str(q.get("list_annual") or "").strip(),
            list_install=str(q.get("list_install") or "").strip(),
        )
        results.append(result)
        if progress_callback is not None:
            try:
                progress_callback(i, total, result)
            except Exception:
                # Progress is best-effort; never break the run if callback fails.
                pass
        if verbose:
            icon = "✓" if code == 0 else "✗"
            print(f"  {icon} {duration:.0f}s")
            if code != 0 and stderr_snippet:
                for line in stderr_snippet.split("\n")[-15:]:
                    print(f"      {line}")
        elif quiet:
            icon = "✓" if code == 0 else "✗"
            print(icon)

    for p in temp_paths:
        try:
            p.unlink(missing_ok=True)
        except Exception:
            pass
    # Optional Demo 2: internal user flow to capture BasketId (kept separate so Demo 1 runtime is unchanged)
    fill_basket_ids_with_internal_user(results, mode)

    # Enforce "full completion" semantics per demo mode.
    # The UI currently keys off `exit_code`/`result` to show green/red.
    # For modes that request BasketId capture, we must fail the row if BasketId isn't captured.
    demo_mode = str(mode).strip().lower()
    require_basket_id = demo_mode in ("demo2", "demo4")
    if require_basket_id:
        for r in results:
            if r.exit_code == 0 and not (getattr(r, "basket_id", "") or "").strip():
                r.exit_code = 1
                r.result = "Fail"
                r.error_detail = (r.error_detail + "\n" if r.error_detail else "") + (
                    "Demo requirement not met: BasketId capture requested, but BasketId was not captured."
                )

    total_duration = time.perf_counter() - start_total
    print_summary(results, total_duration, summary_path, summary_path_powerbi, run_id=run_id_readable, mode=mode)
    failed = sum(1 for r in results if r.exit_code != 0)
    return 1 if failed > 0 else 0, results, summary_path, summary_path_powerbi


def main():
    parser = argparse.ArgumentParser(
        description="Run regression from CSV upload (up to 5 presets). Behind-the-scenes + summary.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 run_csv_regression.py
    → Prompts for CSV path

  python3 run_csv_regression.py example_regression_upload.csv
    → Runs the example CSV

  python3 run_csv_regression.py my.csv --quiet
    → Behind the scenes, summary only (customer-friendly)

  python3 run_csv_regression.py my.csv --verbose
    → Show progress for each preset
        """,
    )
    parser.add_argument("csv", nargs="?", help="Path to CSV file (or omit to be prompted)")
    parser.add_argument("--quiet", action="store_true", help="Behind the scenes: no browser, minimal output, summary at end")
    parser.add_argument("--verbose", "-v", action="store_true", help="Show progress for each preset")
    args = parser.parse_args()

    csv_path: Path | None = None
    if args.csv:
        p = Path(args.csv)
        if not p.is_absolute():
            p = (Path.cwd() / p).resolve()
        if p.exists():
            csv_path = p
        else:
            print(f"  ❌ File not found: {args.csv}")
            sys.exit(1)
    else:
        csv_path = prompt_for_csv()
        if not csv_path:
            print("  No CSV path provided. Exiting.")
            sys.exit(1)

    quiet = args.quiet or not args.verbose  # Default to quiet for customer use
    code, _, _, _ = run_csv_regression(csv_path, quiet=quiet, verbose=args.verbose, mode="demo1")
    sys.exit(code)


if __name__ == "__main__":
    main()
