#!/usr/bin/env python3
"""
Create an Excel template for P2NNI CSV upload with data validation dropdowns.

We support two template variants:
  - Demo 1/2 (customer flow): no negotiated/list annual/install columns
  - Demo 3/4 (internal adjust quote): includes negotiated annual/install + list annual/install

Columns with dropdowns: Product Type, B-End Port, Bandwidth, Term Length,
Shadow VLAN Required, Pay Upfront, FTTP Aggregation, Floor, Room,
Connector Type, Power Supply, Media Type, VLAN Tagging, VLAN Tagging Value, Access Notice,
Auto Negotiation, Hazards on Site, Building Built Prior 2000, Asbestos Register (Yes/No or N/A only — see Instructions),
More Than One Tenant, Land Owner Permission Required.

Columns that stay free text (customer fills in): B-End Postcode, Company Name,
Contact First Name, Contact Surname, Phone, Email, Rack ID, VLAN ID,
PO Reference, Hazards Description.
VLAN Tagging Value is a dropdown: N/A when VLAN Tagging is No; when Yes, same VLAN ID list (1–4094) as Shadow VLAN ID.
(Shadow VLAN ID uses a dropdown: N/A if Shadow VLAN Required is No, else VLAN 1–4094.)

IMPORTANT: Connector Type, Media Type, Auto Negotiation options depend on
Bearer (B-End Port) and Bandwidth. See the "Constraints" sheet for rules.
Validation runs at upload time — invalid combinations will be rejected.

When pasting rows from another Excel: use Paste Special > Values to preserve
dropdown validation. Regular paste overwrites and removes dropdowns.
"""
import csv
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    exit(1)

BASE = Path(__file__).resolve().parent
OPTIONS_CSV = BASE / "csv_dropdown_options.csv"
OUTPUT_XLSX_DEMO12 = BASE / "P2NNI_CSV_Template_Demo12.xlsx"
OUTPUT_XLSX_DEMO34 = BASE / "P2NNI_CSV_Template_Demo34.xlsx"

# Headers: order matches digital portal UI (Floor, Room, Rack ID; site config + readiness; then VLAN IDs; PO Ref last)
HEADERS_FULL = [
    "Product Type", "B-End Postcode", "B-End Port (Mbps)", "Bandwidth (Mbps)",
    "Term Length (months)", "Shadow VLAN Required", "Pay Upfront", "FTTP Aggregation",
    "Preferred Supplier",
    "Negotiated Annual", "Negotiated Install",
    "List Annual", "List Install",
    "Company Name", "Contact First Name", "Contact Surname",
    "Phone", "Email", "Floor", "Room", "Rack ID",
    "Connector Type", "Power Supply", "Media Type", "VLAN Tagging", "VLAN Tagging Value",
    "Access Notice", "Auto Negotiation", "Hazards on Site", "Hazards Description",
    "Building Built Prior 2000", "Asbestos Register", "More Than One Tenant",
    "Land Owner Permission Required", "VLAN ID", "Shadow VLAN ID", "PO Reference",
]

# Example row (can be overwritten by user)
EXAMPLE_ROW_FULL = [
    "P2NNI", "SP2 8NJ", "1000", "1000", "36", "Yes", "Yes", "No",
    "",  # Preferred Supplier (optional; blank = choose cheapest)
    "", "",  # Negotiated Annual, Negotiated Install (Demo 3/4; optional, e.g. 4054.05)
    "", "",  # List Annual, List Install (optional; if blank, scraped from quote page for discount)
    "Test Co", "Test", "User", "07123456789", "test@example.com",
    "001 - 1st Floor", "ADMN - Admin Room", "1",
    "LC", "AC", "LR", "No", "N/A", "0-48 hours", "Yes", "Yes", "Standard building hazards",
    "No", "N/A", "No", "No", "100", "100", "TEST-PO-001",
]

NEGOTIATED_LIST_COLS = {"Negotiated Annual", "Negotiated Install", "List Annual", "List Install"}
HEADERS_DEMO12 = [h for h in HEADERS_FULL if h not in NEGOTIATED_LIST_COLS]
EXAMPLE_ROW_DEMO12 = [v for h, v in zip(HEADERS_FULL, EXAMPLE_ROW_FULL) if h not in NEGOTIATED_LIST_COLS]


def load_dropdown_options() -> dict[str, list[str]]:
    """Load column -> [options] from csv_dropdown_options.csv."""
    if not OPTIONS_CSV.exists():
        return {}
    options: dict[str, list[str]] = {}
    with open(OPTIONS_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            col = row.get("Column", "").strip()
            opt = row.get("Option", "").strip()
            if not col or not opt:
                continue
            if col not in options:
                options[col] = []
            options[col].append(opt)
    return options


def create_template(demo_mode: str = "demo34"):
    """
    Create one Excel template.

    demo_mode:
      - "demo12" -> Demo 1/2 template (no negotiated/list annual/install columns)
      - "demo34" -> Demo 3/4 template (includes negotiated/list annual/install columns)
    """
    demo_mode = str(demo_mode or "").strip().lower()
    if demo_mode not in ("demo12", "demo34"):
        demo_mode = "demo34"

    include_negotiated = demo_mode == "demo34"
    headers = HEADERS_FULL if include_negotiated else HEADERS_DEMO12
    example_row = EXAMPLE_ROW_FULL if include_negotiated else EXAMPLE_ROW_DEMO12
    output_xlsx = OUTPUT_XLSX_DEMO34 if include_negotiated else OUTPUT_XLSX_DEMO12

    wb = Workbook()
    # Instructions sheet first (critical for paste workflow)
    inst = wb.create_sheet("Instructions", 0)
    inst.column_dimensions["A"].width = 80
    inst.cell(row=1, column=1, value="P2NNI CSV Formatting Guide – Important")
    inst.cell(row=2, column=1, value="")
    inst.cell(row=3, column=1, value="WHEN PASTING FROM ANOTHER EXCEL:")
    inst.cell(row=4, column=1, value="  • Do NOT use regular Paste (Ctrl+V / Cmd+V) – it removes dropdowns")
    inst.cell(row=5, column=1, value="  • Use: Edit → Paste Special → Values (or Right‑click → Paste Special → Values)")
    inst.cell(row=6, column=1, value="  • Paste Special → Values keeps the dropdown lists on pasted cells")
    inst.cell(row=7, column=1, value="")
    inst.cell(row=8, column=1, value="WORKFLOW: Keep row 1 (headers). Paste your data into row 2 onwards using Paste Special → Values.")
    inst.cell(row=9, column=1, value="BEARER / BANDWIDTH: B-End Port (bearer) restricts Bandwidth options:")
    inst.cell(row=10, column=1, value="  • 10 Gbps bearer → Bandwidth: 10000 or 1000 only")
    inst.cell(row=11, column=1, value="  • 1 Gbps bearer → Bandwidth: 1000, 500, 200, or 100")
    inst.cell(row=12, column=1, value="  • 100 Mbps bearer → Bandwidth: 100 only")
    inst.cell(row=13, column=1, value="Connector, Media, Auto Negotiation, etc. then depend on Bearer+Bandwidth — see Constraints sheet.")
    inst.cell(
        row=14,
        column=1,
        value=(
            "VLAN Tagging Value: when \"VLAN Tagging\" is No, the dropdown is N/A only; when Yes, choose a B-End tag (1–4094), or leave blank to reuse \"VLAN ID\" in CSV/automation. "
            "If Shadow VLAN Required = No, the Shadow VLAN ID dropdown is N/A only; if Yes, choose a VLAN ID (1–4094). "
            "Automation applies N/A on the portal when shadow is not required or VLAN tagging is off. "
            "If Building Pre-2000 = No, the Asbestos Register dropdown is N/A only (Excel enforces this per row)."
        ),
    )
    if include_negotiated:
        inst.cell(
            row=15,
            column=1,
            value=(
                "AFTER PREFERRED SUPPLIER (Demo 3/4 pricing): Negotiated Annual, Negotiated Install (target prices). "
                "Optional: List Annual, List Install (if blank, list prices are taken from the quote page for discount = list − negotiated)."
            ),
        )
    else:
        inst.cell(
            row=15,
            column=1,
            value="DEMO 1/2 TEMPLATE: negotiated/list annual/install columns are intentionally excluded (customer flow has no adjust-quote pricing inputs).",
        )
    inst.cell(row=16, column=1, value="SAVE AS CSV: Use File → Save As → 'CSV UTF-8 (Comma delimited)(*.csv)' so columns stay aligned. If B-End Postcode is only in row 2, leave it blank in later rows — the first row's postcode is used for all.")
    inst.cell(row=17, column=1, value="")

    ws = wb.worksheets[1]
    ws.title = "P2NNI Upload"
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    # Write example row (row 2)
    for col_idx, val in enumerate(example_row, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    # Set column widths so headers and values fit cleanly (Excel units ≈ character width)
    for i in range(len(headers)):
        header_len = len(str(headers[i]))
        example_len = len(str(example_row[i])) if i < len(example_row) else 0
        w = min(max(header_len, example_len) + 2, 45)
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Column name -> 1-based index
    col_map = {h: i for i, h in enumerate(headers, start=1)}
    options = load_dropdown_options()

    # Dynamic columns: options depend on B-End Port (col C) and Bandwidth (col D)
    DYNAMIC_COLS = ("Connector Type", "Power Supply", "Media Type", "VLAN Tagging", "Auto Negotiation")
    bearer_col = col_map.get("B-End Port (Mbps)", 3)  # C
    bw_col = col_map.get("Bandwidth (Mbps)", 4)  # D

    BANDWIDTH_BY_BEARER = {
        10000: [10000, 1000],
        1000: [1000, 500, 200, 100],
        100: [100],
    }
    bandwidth_dynamic = False
    # Constraints reference sheet + dynamic option sheets with named ranges
    try:
        from p2nni_constraints import _CONSTRAINTS
        # Media type: Excel shows portal labels. 10G bearer uses LR; 1G/100M use LX/SX.
        def _media_display(media_code: str, bearer_mbps: int) -> str:
            if media_code == "LR":
                return "Single Mode (LR)" if bearer_mbps >= 10000 else "Single Mode (LX)"
            if media_code == "SR":
                return "Multi Mode (SX)"
            return media_code
        c_ws = wb.create_sheet("Constraints")
        c_ws.cell(row=1, column=1, value="Bearer/Bandwidth")
        c_ws.cell(row=1, column=2, value="Connector Type")
        c_ws.cell(row=1, column=3, value="Media Type")
        c_ws.cell(row=1, column=4, value="Power Supply")
        c_ws.cell(row=1, column=5, value="Auto Negotiation")
        c_ws.cell(row=1, column=6, value="VLAN Tagging")
        labels = {
            (10000, 10000): "10 Gbps / 10 Gbps",
            (10000, 1000): "10 Gbps / 1 Gbps",
            (1000, 1000): "1 Gbps / 1 Gbps",
            (1000, 500): "1 Gbps / 500 Mbps",
            (1000, 200): "1 Gbps / 200 Mbps",
            (1000, 100): "1 Gbps / 100 Mbps",
            (100, 100): "100 Mbps / 100 Mbps",
        }
        for r, (key, cfg) in enumerate(_CONSTRAINTS.items(), start=2):
            c_ws.cell(row=r, column=1, value=labels.get(key, f"{key[0]}/{key[1]}"))
            c_ws.cell(row=r, column=2, value=", ".join(cfg["connector_type"]))
            media_display = [_media_display(m, key[0]) for m in cfg["media_type"]]
            c_ws.cell(row=r, column=3, value=", ".join(media_display))
            c_ws.cell(row=r, column=4, value=", ".join(cfg["power_supply"]))
            c_ws.cell(row=r, column=5, value=", ".join(cfg["auto_negotiation"]))
            c_ws.cell(row=r, column=6, value=", ".join(cfg["vlan_tagging"]))
        for c in range(1, 7):
            c_ws.column_dimensions[get_column_letter(c)].width = 22
        # Bandwidth by Bearer (for reference)
        c_ws.cell(row=10, column=1, value="Bandwidth by Bearer")
        c_ws.cell(row=11, column=1, value="10 Gbps bearer")
        c_ws.cell(row=11, column=2, value="10 Gbps, 1 Gbps")
        c_ws.cell(row=12, column=1, value="1 Gbps bearer")
        c_ws.cell(row=12, column=2, value="1 Gbps, 500, 200, 100 Mbps")
        c_ws.cell(row=13, column=1, value="100 Mbps bearer")
        c_ws.cell(row=13, column=2, value="100 Mbps only")

        # Create sheets with option columns per bearer/bandwidth key; define named ranges
        field_to_sheet = {
            "connector_type": "DynConnector",
            "power_supply": "DynPower",
            "media_type": "DynMedia",
            "auto_negotiation": "DynAutoNeg",
            "vlan_tagging": "DynVlan",
        }
        keys_ordered = list(_CONSTRAINTS.keys())
        for field, sheet_name in field_to_sheet.items():
            dyn_ws = wb.create_sheet(sheet_name)
            for c, key in enumerate(keys_ordered, start=1):
                opts = _CONSTRAINTS[key][field]
                if field == "media_type":
                    opts = [_media_display(o, key[0]) for o in opts]
                for r, o in enumerate(opts, start=2):
                    dyn_ws.cell(row=r, column=c, value=o)
                end_row = 1 + len(opts)
                name = f"{sheet_name}_{key[0]}_{key[1]}"
                ref = f"'{sheet_name}'!${get_column_letter(c)}$2:${get_column_letter(c)}${end_row}"
                wb.defined_names[name] = DefinedName(name, attr_text=ref)

        # Bandwidth depends on Bearer only: 10G→10 or 1; 1G→1,500,200,100; 100M→100
        dyn_bw = wb.create_sheet("DynBandwidth")
        for c, (bearer, bw_opts) in enumerate(BANDWIDTH_BY_BEARER.items(), start=1):
            for r, val in enumerate(bw_opts, start=2):
                dyn_bw.cell(row=r, column=c, value=str(val))
            name = f"DynBandwidth_{bearer}"
            ref = f"'DynBandwidth'!${get_column_letter(c)}$2:${get_column_letter(c)}${1 + len(bw_opts)}"
            wb.defined_names[name] = DefinedName(name, attr_text=ref)

        # Bandwidth dropdown: INDIRECT so 10G bearer → only 10 Gbps & 1 Gbps; 1G → 1,500,200,100; 100M → 100
        bear_letter = get_column_letter(bearer_col)
        bw_main = get_column_letter(bw_col)
        for row in range(2, 201):
            formula = f'=INDIRECT("DynBandwidth_"&{bear_letter}{row})'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
            dv.error = "Choose Bandwidth allowed for this Bearer. 10G→10 or 1; 1G→1,500,200,100; 100M→100."
            dv.errorTitle = "Invalid entry"
            ws.add_data_validation(dv)
            dv.add(f"{bw_main}{row}")
        bandwidth_dynamic = True
    except ImportError:
        bandwidth_dynamic = False

    # Options sheet: one column per dropdown. Use range refs for static columns only.
    opts_sheet_name = "Options"
    opts_ws = wb.create_sheet(opts_sheet_name)
    opts_ws.sheet_state = "visible"  # Hidden sheets can cause Excel to "repair" and drop dropdowns
    opts_col = 0

    for col_name, opts in options.items():
        col_idx = col_map.get(col_name)
        if col_idx is None or not opts:
            continue
        if col_name in DYNAMIC_COLS:
            continue  # Handled separately with INDIRECT
        if col_name == "Bandwidth (Mbps)" and bandwidth_dynamic:
            continue  # Handled by DynBandwidth INDIRECT
        if col_name == "Asbestos Register":
            continue  # Depends on Building Built Prior 2000 — INDIRECT + named ranges below
        opts_col += 1
        col_letter = get_column_letter(opts_col)
        for r, o in enumerate(opts, start=1):
            opts_ws.cell(row=r, column=opts_col, value=o)
        # Cross-sheet ref: =Options!$A$1:$A$5
        formula = f"={opts_sheet_name}!${col_letter}$1:${col_letter}${len(opts)}"
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=False,  # False = show dropdown (openpyxl: inverted naming)
        )
        dv.error = "Choose from the dropdown list"
        dv.errorTitle = "Invalid entry"
        ws.add_data_validation(dv)
        main_col = get_column_letter(col_idx)
        dv.add(f"{main_col}2:{main_col}200")

    # Fallback: static Bandwidth when DynBandwidth not created (e.g. p2nni_constraints missing)
    if not bandwidth_dynamic:
        opts = options.get("Bandwidth (Mbps)", ["100", "200", "500", "1000", "10000"])
        if opts:
            opts_col += 1
            col_letter = get_column_letter(opts_col)
            for r, o in enumerate(opts, start=1):
                opts_ws.cell(row=r, column=opts_col, value=o)
            formula = f"={opts_sheet_name}!${col_letter}$1:${col_letter}${len(opts)}"
            dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
            dv.error = "Choose from the dropdown list"
            dv.errorTitle = "Invalid entry"
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(bw_col)}2:{get_column_letter(bw_col)}200")

    # Dynamic dropdowns: Connector, Media, Power, VLAN, Auto Negotiation (INDIRECT per row)
    dyn_sheet_prefix = {"Connector Type": "DynConnector", "Power Supply": "DynPower", "Media Type": "DynMedia",
                        "VLAN Tagging": "DynVlan", "Auto Negotiation": "DynAutoNeg"}
    try:
        from p2nni_constraints import _CONSTRAINTS
        bear_letter = get_column_letter(bearer_col)
        bw_letter = get_column_letter(bw_col)
        for col_name in DYNAMIC_COLS:
            col_idx = col_map.get(col_name)
            if col_idx is None:
                continue
            prefix = dyn_sheet_prefix.get(col_name, "")
            if not prefix:
                continue
            main_col = get_column_letter(col_idx)
            for row in range(2, 201):
                formula = f'=INDIRECT("{prefix}_"&{bear_letter}{row}&"_"&{bw_letter}{row})'
                dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
                dv.error = f"Choose from options allowed for this Bearer/Bandwidth. See Constraints sheet."
                dv.errorTitle = "Invalid entry"
                ws.add_data_validation(dv)
                dv.add(f"{main_col}{row}")
    except ImportError:
        # Fallback: use static options for dynamic columns if constraints module missing
        for col_name in DYNAMIC_COLS:
            opts = options.get(col_name, [])
            col_idx = col_map.get(col_name)
            if col_idx is None or not opts:
                continue
            opts_col += 1
            col_letter = get_column_letter(opts_col)
            for r, o in enumerate(opts, start=1):
                opts_ws.cell(row=r, column=opts_col, value=o)
            formula = f"={opts_sheet_name}!${col_letter}$1:${col_letter}${len(opts)}"
            dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
            dv.error = "Choose from the dropdown list"
            dv.errorTitle = "Invalid entry"
            ws.add_data_validation(dv)
            main_col = get_column_letter(col_idx)
            dv.add(f"{main_col}2:{main_col}200")

    # Shared Options! cell for "N/A" lists (asbestos + shadow VLAN ID when not required)
    ref_na_shared = None  # Options! single cell for "N/A" (asbestos and/or shadow VLAN ID)

    # Asbestos Register: Building Built Prior 2000 = No → dropdown only N/A; = Yes → Yes or No
    b_prior_idx = col_map.get("Building Built Prior 2000")
    asb_idx = col_map.get("Asbestos Register")
    if b_prior_idx and asb_idx:
        opts_col += 1
        c_yes_no = get_column_letter(opts_col)
        opts_ws.cell(row=1, column=opts_col, value="Yes")
        opts_ws.cell(row=2, column=opts_col, value="No")
        opts_col += 1
        c_na = get_column_letter(opts_col)
        opts_ws.cell(row=1, column=opts_col, value="N/A")
        ref_yes_no = f"'{opts_sheet_name}'!${c_yes_no}$1:${c_yes_no}$2"
        ref_na_only = f"'{opts_sheet_name}'!${c_na}$1:${c_na}$1"
        wb.defined_names["Asbestos_When_Pre2000"] = DefinedName(
            "Asbestos_When_Pre2000", attr_text=ref_yes_no
        )
        wb.defined_names["Asbestos_NA_Only"] = DefinedName("Asbestos_NA_Only", attr_text=ref_na_only)
        ref_na_shared = ref_na_only
        bp_letter = get_column_letter(b_prior_idx)
        asb_letter = get_column_letter(asb_idx)
        for row in range(2, 201):
            formula = f'=INDIRECT(IF({bp_letter}{row}="No","Asbestos_NA_Only","Asbestos_When_Pre2000"))'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
            dv.error = (
                'When "Building Built Prior 2000" is No, Asbestos Register must be N/A. '
                "When Yes, choose Yes or No."
            )
            dv.errorTitle = "Invalid entry"
            ws.add_data_validation(dv)
            dv.add(f"{asb_letter}{row}")

    # Shadow VLAN ID: Shadow VLAN Required = No → N/A only; = Yes → VLAN IDs 1–4094 (portal range)
    sreq_idx = col_map.get("Shadow VLAN Required")
    svid_idx = col_map.get("Shadow VLAN ID")
    if sreq_idx and svid_idx:
        if ref_na_shared is None:
            opts_col += 1
            c_na_sv = get_column_letter(opts_col)
            opts_ws.cell(row=1, column=opts_col, value="N/A")
            ref_na_shared = f"'{opts_sheet_name}'!${c_na_sv}$1:${c_na_sv}$1"
        wb.defined_names["ShadowVlanId_NA_Only"] = DefinedName(
            "ShadowVlanId_NA_Only", attr_text=ref_na_shared
        )
        svl = wb.create_sheet("ShadowVlanIdList")
        svl.sheet_state = "hidden"
        for i in range(1, 4095):
            svl.cell(row=i, column=1, value=i)
        ref_vlan_ids = "'ShadowVlanIdList'!$A$1:$A$4094"
        wb.defined_names["ShadowVlanId_WhenRequired"] = DefinedName(
            "ShadowVlanId_WhenRequired", attr_text=ref_vlan_ids
        )
        sreq_letter = get_column_letter(sreq_idx)
        svid_letter = get_column_letter(svid_idx)
        for row in range(2, 201):
            formula = f'=INDIRECT(IF({sreq_letter}{row}="No","ShadowVlanId_NA_Only","ShadowVlanId_WhenRequired"))'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
            dv.error = (
                'When "Shadow VLAN Required" is No, Shadow VLAN ID must be N/A. '
                "When Yes, pick a VLAN ID from the list (1–4094)."
            )
            dv.errorTitle = "Invalid entry"
            ws.add_data_validation(dv)
            dv.add(f"{svid_letter}{row}")

    # VLAN Tagging Value: VLAN Tagging = No → N/A only; = Yes → same 1–4094 list as Shadow VLAN (B-End tag)
    vt_idx = col_map.get("VLAN Tagging")
    vtv_idx = col_map.get("VLAN Tagging Value")
    if vt_idx and vtv_idx and "ShadowVlanId_WhenRequired" in wb.defined_names:
        if ref_na_shared is None:
            opts_col += 1
            c_na_vtv = get_column_letter(opts_col)
            opts_ws.cell(row=1, column=opts_col, value="N/A")
            ref_na_shared = f"'{opts_sheet_name}'!${c_na_vtv}$1:${c_na_vtv}$1"
        wb.defined_names["VlanTaggingValue_NA_Only"] = DefinedName(
            "VlanTaggingValue_NA_Only", attr_text=ref_na_shared
        )
        vt_letter = get_column_letter(vt_idx)
        vtv_letter = get_column_letter(vtv_idx)
        for row in range(2, 201):
            formula = f'=INDIRECT(IF({vt_letter}{row}="No","VlanTaggingValue_NA_Only","ShadowVlanId_WhenRequired"))'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
            dv.error = (
                'When "VLAN Tagging" is No, VLAN Tagging Value must be N/A. '
                "When Yes, pick a VLAN ID from the list (1–4094)."
            )
            dv.errorTitle = "Invalid entry"
            ws.add_data_validation(dv)
            dv.add(f"{vtv_letter}{row}")

    # Add paste tip to Constraints sheet
    try:
        c_ws.cell(row=14, column=1, value="TIP: When pasting from another Excel, use Paste Special > Values to keep dropdowns.")
    except Exception:
        pass

    wb.save(output_xlsx)
    print(f"Created {output_xlsx}")
    print("Open in Excel, fill your rows, then save as CSV for upload.")
    print("TIP: When pasting from another Excel, use Paste Special > Values to preserve dropdowns.")


if __name__ == "__main__":
    create_template("demo34")
